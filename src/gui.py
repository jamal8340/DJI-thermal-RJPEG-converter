import csv
import os
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from converter import (
    convert_folder,
    convert_images,
    create_sdk,
)

from validator import (
    validate_files,
)


def resource_path(relative_path):
    """Return an absolute path for development and PyInstaller onefile builds."""
    if getattr(sys, "frozen", False):
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path(__file__).resolve().parent.parent

    return base_path / relative_path


BG = "#F7F1E7"
CARD_BG = "#FFFDF8"
TEXT = "#171717"
MUTED = "#6D675F"
ACCENT = "#D39A2C"
ACCENT_HOVER = "#B9801E"
BORDER = "#DDD5C8"


class ConverterGUI:
    def __init__(self, root):
        self.root = root

        # Window/taskbar icon. Keep a PhotoImage reference on self so Tkinter
        # does not garbage-collect it. The PNG is bundled by PyInstaller.
        try:
            icon_path = resource_path("assets/app_icon.png")
            self.app_icon = tk.PhotoImage(file=str(icon_path))
            self.root.iconphoto(True, self.app_icon)
        except Exception as exc:
            # The application must remain usable even if the icon cannot load.
            print(f"Could not load window icon: {exc}")

        self.root.title(
            "DJI Thermal R-JPEG Converter v1.0.0"
        )

        # Comfortable default size: the complete workflow is visible at launch.
        self.root.geometry("1120x740")
        self.root.minsize(1020, 700)

        self.root.configure(
            bg=BG
        )

        if getattr(sys, "frozen", False):
            self.default_output = (
                Path.home()
                / "Documents"
                / "DJI_Thermal_Converter"
                / "output"
            )
        else:
            base_dir = (
                Path(__file__)
                .resolve()
                .parent
                .parent
            )

            self.default_output = (
                base_dir
                / "data"
                / "output"
            )

        self.selected_files = []
        self.input_folder = None
        self.selection_mode = None

        self.last_output_dir = None
        self.output_manually_selected = False

        self.output_path = tk.StringVar(
            value=str(
                self.default_output
            )
        )

        self.input_status = tk.StringVar(
            value=""
        )

        self.output_status = tk.StringVar(
            value=(
                "✓ Default output folder selected"
            )
        )

        self.main_status = tk.StringVar(
            value="Ready"
        )

        self.overwrite_existing = (
            tk.BooleanVar(
                value=False
            )
        )

        self.use_auto_output = tk.BooleanVar(
            value=True
        )

        self.use_image_radiometry = tk.BooleanVar(
            value=True
        )

        self.emissivity_value = tk.StringVar(
            value=""
        )

        self.distance_value = tk.StringVar(
            value=""
        )

        self.humidity_value = tk.StringVar(
            value=""
        )

        self.reflection_value = tk.StringVar(
            value=""
        )

        self.source_radiometry = None

        self.setup_style()
        self.create_widgets()

    def setup_style(self):
        style = ttk.Style()

        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        # Base surfaces
        style.configure(
            "Main.TFrame",
            background=BG
        )

        style.configure(
            "Card.TFrame",
            background=CARD_BG
        )

        style.configure(
            "Path.TFrame",
            background="#FBF7F0"
        )

        style.configure(
            "Action.TFrame",
            background="#F1E7D6"
        )

        # Typography
        style.configure(
            "Title.TLabel",
            background=BG,
            foreground=TEXT,
            font=("Segoe UI", 23, "bold")
        )

        style.configure(
            "Subtitle.TLabel",
            background=BG,
            foreground=MUTED,
            font=("Segoe UI", 10)
        )

        style.configure(
            "Version.TLabel",
            background="#EFE3CE",
            foreground=ACCENT_HOVER,
            font=("Segoe UI", 9, "bold"),
            padding=(10, 5)
        )

        style.configure(
            "Section.TLabel",
            background=CARD_BG,
            foreground=TEXT,
            font=("Segoe UI", 11, "bold")
        )

        style.configure(
            "SectionHint.TLabel",
            background=CARD_BG,
            foreground=MUTED,
            font=("Segoe UI", 9)
        )

        style.configure(
            "Muted.TLabel",
            background=CARD_BG,
            foreground=MUTED,
            font=("Segoe UI", 9)
        )

        style.configure(
            "Path.TLabel",
            background="#FBF7F0",
            foreground="#4F4A43",
            font=("Segoe UI", 9)
        )

        style.configure(
            "Success.TLabel",
            background=CARD_BG,
            foreground=ACCENT_HOVER,
            font=("Segoe UI", 9, "bold")
        )

        style.configure(
            "Status.TLabel",
            background="#F1E7D6",
            foreground=MUTED,
            font=("Segoe UI", 9)
        )

        style.configure(
            "Percent.TLabel",
            background="#F1E7D6",
            foreground=TEXT,
            font=("Segoe UI", 10, "bold")
        )

        # Buttons
        style.configure(
            "Secondary.TButton",
            background="#FFFDF8",
            foreground=TEXT,
            bordercolor=BORDER,
            lightcolor=BORDER,
            darkcolor=BORDER,
            borderwidth=1,
            padding=(14, 8),
            font=("Segoe UI", 9, "bold")
        )

        style.map(
            "Secondary.TButton",
            background=[
                ("active", "#F4ECDF"),
                ("disabled", "#EEEAE3"),
            ],
            foreground=[
                ("disabled", "#AAA39A"),
            ],
            bordercolor=[
                ("active", "#CDBD9F"),
            ]
        )

        style.configure(
            "Accent.TButton",
            background=ACCENT,
            foreground="white",
            bordercolor=ACCENT,
            lightcolor=ACCENT,
            darkcolor=ACCENT,
            borderwidth=0,
            padding=(30, 12),
            font=("Segoe UI", 10, "bold")
        )

        style.map(
            "Accent.TButton",
            background=[
                ("active", ACCENT_HOVER),
                ("disabled", "#C9B88D"),
            ],
            foreground=[
                ("disabled", "#F4EFE7"),
            ]
        )

        # Inputs
        style.configure(
            "Field.TEntry",
            fieldbackground="#FFFFFF",
            foreground=TEXT,
            insertcolor=TEXT,
            bordercolor=BORDER,
            lightcolor=BORDER,
            darkcolor=BORDER,
            borderwidth=1,
            padding=(8, 7)
        )

        style.map(
            "Field.TEntry",
            fieldbackground=[
                ("disabled", "#F2EEE7"),
                ("readonly", "#F2EEE7"),
            ],
            foreground=[
                ("disabled", "#777168"),
            ],
            bordercolor=[
                ("focus", ACCENT),
            ]
        )

        style.configure(
            "Custom.TCheckbutton",
            background=CARD_BG,
            foreground=MUTED,
            font=("Segoe UI", 9),
            padding=0
        )

        style.map(
            "Custom.TCheckbutton",
            background=[
                ("active", CARD_BG),
            ],
            foreground=[
                ("active", TEXT),
            ]
        )

        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor="#E4D9C8",
            background=ACCENT,
            bordercolor="#E4D9C8",
            lightcolor=ACCENT,
            darkcolor=ACCENT,
            thickness=10
        )

    def create_card(self, parent):
        card = tk.Frame(
            parent,
            bg=CARD_BG,
            highlightbackground=BORDER,
            highlightthickness=1,
            bd=0
        )

        inner = ttk.Frame(
            card,
            style="Card.TFrame",
            padding=(20, 18)
        )

        inner.pack(
            fill="both",
            expand=True
        )

        return card, inner

    def create_section_header(self, parent, title, hint):
        ttk.Label(
            parent,
            text=title,
            style="Section.TLabel"
        ).pack(
            anchor="w"
        )

        ttk.Label(
            parent,
            text=hint,
            style="SectionHint.TLabel"
        ).pack(
            anchor="w",
            pady=(3, 12)
        )

    def create_widgets(self):
        container = ttk.Frame(
            self.root,
            style="Main.TFrame",
            padding=(28, 24)
        )

        container.pack(
            fill="both",
            expand=True
        )

        # HEADER
        header = ttk.Frame(
            container,
            style="Main.TFrame"
        )
        header.pack(
            fill="x",
            pady=(0, 20)
        )

        title_row = ttk.Frame(
            header,
            style="Main.TFrame"
        )
        title_row.pack(
            fill="x"
        )

        title_group = ttk.Frame(
            title_row,
            style="Main.TFrame"
        )
        title_group.pack(
            side="left",
            fill="x",
            expand=True
        )

        # A restrained gold accent gives the header a stronger product identity.
        accent_bar = tk.Frame(
            title_group,
            bg=ACCENT,
            width=4,
            height=48
        )
        accent_bar.pack(
            side="left",
            padx=(0, 14)
        )
        accent_bar.pack_propagate(False)

        title_text = ttk.Frame(
            title_group,
            style="Main.TFrame"
        )
        title_text.pack(
            side="left"
        )

        ttk.Label(
            title_text,
            text="DJI Thermal R-JPEG Converter v1.0.0",
            style="Title.TLabel"
        ).pack(
            anchor="w"
        )

        ttk.Label(
            title_text,
            text=(
                "Radiometric DJI imagery → compressed Float32 "
                "temperature TIFF"
            ),
            style="Subtitle.TLabel"
        ).pack(
            anchor="w",
            pady=(3, 0)
        )

        # INPUT + OUTPUT: use the width instead of stacking two large empty cards.
        io_grid = ttk.Frame(
            container,
            style="Main.TFrame"
        )
        io_grid.pack(
            fill="x",
            pady=(0, 14)
        )

        io_grid.columnconfigure(0, weight=1, uniform="io")
        io_grid.columnconfigure(1, weight=1, uniform="io")

        input_outer, input_card = self.create_card(io_grid)
        input_outer.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(0, 7)
        )

        output_outer, output_card = self.create_card(io_grid)
        output_outer.grid(
            row=0,
            column=1,
            sticky="nsew",
            padx=(7, 0)
        )

        # INPUT
        self.create_section_header(
            input_card,
            "Input images",
            "Choose individual DJI R-JPEG files or an entire folder."
        )

        input_path_box = tk.Frame(
            input_card,
            bg="#FBF7F0",
            highlightbackground="#E8E0D4",
            highlightthickness=1,
            bd=0
        )
        input_path_box.pack(
            fill="x",
            pady=(0, 12)
        )

        self.input_path_label = ttk.Label(
            input_path_box,
            text="No input selected",
            style="Path.TLabel",
            wraplength=430,
            justify="left"
        )
        self.input_path_label.pack(
            fill="x",
            padx=11,
            pady=9
        )

        input_row = ttk.Frame(
            input_card,
            style="Card.TFrame"
        )
        input_row.pack(
            fill="x"
        )

        ttk.Button(
            input_row,
            text="Select images",
            style="Secondary.TButton",
            command=self.select_images
        ).pack(
            side="left",
            padx=(0, 8)
        )

        ttk.Button(
            input_row,
            text="Select folder",
            style="Secondary.TButton",
            command=self.select_folder
        ).pack(
            side="left"
        )

        ttk.Checkbutton(
            input_card,
            text="Overwrite existing TIFF files",
            variable=self.overwrite_existing,
            style="Custom.TCheckbutton"
        ).pack(
            anchor="w",
            pady=(12, 0)
        )

        self.input_status_label = ttk.Label(
            input_card,
            textvariable=self.input_status,
            style="Muted.TLabel"
        )
        self.input_status_label.pack(
            anchor="w",
            pady=(8, 0)
        )

        # OUTPUT
        self.create_section_header(
            output_card,
            "Output folder",
            "Use the suggested folder or switch to a custom location."
        )

        output_path_box = tk.Frame(
            output_card,
            bg="#FBF7F0",
            highlightbackground="#E8E0D4",
            highlightthickness=1,
            bd=0
        )
        output_path_box.pack(
            fill="x",
            pady=(0, 12)
        )

        ttk.Label(
            output_path_box,
            textvariable=self.output_path,
            style="Path.TLabel",
            wraplength=430,
            justify="left"
        ).pack(
            fill="x",
            padx=11,
            pady=9
        )

        ttk.Checkbutton(
            output_card,
            text="Use automatic output folder",
            variable=self.use_auto_output,
            style="Custom.TCheckbutton",
            command=self.toggle_output_mode
        ).pack(
            anchor="w",
            pady=(0, 12)
        )

        output_buttons = ttk.Frame(
            output_card,
            style="Card.TFrame"
        )
        output_buttons.pack(
            fill="x"
        )

        self.change_output_button = ttk.Button(
            output_buttons,
            text="Change output folder",
            style="Secondary.TButton",
            command=self.select_output,
            state="disabled"
        )
        self.change_output_button.pack(
            side="left",
            padx=(0, 8)
        )

        self.output_status_label = ttk.Label(
            output_card,
            textvariable=self.output_status,
            style="Success.TLabel"
        )
        self.output_status_label.pack(
            anchor="w",
            pady=(12, 0)
        )

        # RADIOMETRIC PARAMETERS
        radiometry_outer, radiometry_card = self.create_card(container)
        radiometry_outer.pack(
            fill="x",
            pady=(0, 14)
        )

        radiometry_header = ttk.Frame(
            radiometry_card,
            style="Card.TFrame"
        )
        radiometry_header.pack(
            fill="x"
        )

        radiometry_title = ttk.Frame(
            radiometry_header,
            style="Card.TFrame"
        )
        radiometry_title.pack(
            side="left",
            fill="x",
            expand=True
        )

        ttk.Label(
            radiometry_title,
            text="Radiometric parameters",
            style="Section.TLabel"
        ).pack(
            anchor="w"
        )

        ttk.Label(
            radiometry_title,
            text="Control how the DJI SDK calculates the temperature raster.",
            style="SectionHint.TLabel"
        ).pack(
            anchor="w",
            pady=(3, 0)
        )

        ttk.Checkbutton(
            radiometry_header,
            text="Use values stored in each image",
            variable=self.use_image_radiometry,
            style="Custom.TCheckbutton",
            command=self.toggle_radiometry_fields
        ).pack(
            side="right",
            anchor="n",
            pady=(2, 0)
        )

        radiometry_grid = ttk.Frame(
            radiometry_card,
            style="Card.TFrame"
        )
        radiometry_grid.pack(
            fill="x",
            pady=(16, 0)
        )

        labels = [
            (
                "Emissivity",
                "0.10–1.00",
                self.emissivity_value
            ),
            (
                "Distance",
                "m · 1–25",
                self.distance_value
            ),
            (
                "Humidity",
                "% · 1–100",
                self.humidity_value
            ),
            (
                "Reflected temperature",
                "°C · -40–100",
                self.reflection_value
            ),
        ]

        self.radiometry_entries = []

        for column, (label_text, range_text, variable) in enumerate(labels):
            field = ttk.Frame(
                radiometry_grid,
                style="Card.TFrame"
            )

            field.grid(
                row=0,
                column=column,
                sticky="ew",
                padx=(0, 12 if column < 3 else 0)
            )

            radiometry_grid.columnconfigure(
                column,
                weight=1,
                uniform="radiometry"
            )

            label_row = ttk.Frame(
                field,
                style="Card.TFrame"
            )
            label_row.pack(
                fill="x"
            )

            ttk.Label(
                label_row,
                text=label_text,
                style="SectionHint.TLabel"
            ).pack(
                side="left"
            )

            ttk.Label(
                label_row,
                text=range_text,
                style="Muted.TLabel"
            ).pack(
                side="right"
            )

            entry = ttk.Entry(
                field,
                textvariable=variable,
                style="Field.TEntry"
            )

            entry.bind(
                "<FocusOut>",
                lambda event, var=variable: (
                    self.on_radiometry_focus_out(var)
                )
            )

            entry.bind(
                "<Return>",
                lambda event, var=variable: (
                    self.on_radiometry_focus_out(var)
                )
            )

            entry.pack(
                fill="x",
                pady=(6, 0)
            )

            self.radiometry_entries.append(entry)

        self.toggle_radiometry_fields()

        # CONVERSION / PROGRESS
        action_outer = tk.Frame(
            container,
            bg="#F1E7D6",
            highlightbackground="#DCC9AA",
            highlightthickness=1,
            bd=0
        )
        action_outer.pack(
            fill="x"
        )

        action = ttk.Frame(
            action_outer,
            style="Action.TFrame",
            padding=(20, 16)
        )
        action.pack(
            fill="both",
            expand=True
        )

        # Main conversion action. Keep the CTA visually central in the workflow.
        self.convert_button = ttk.Button(
            action,
            text="Convert images",
            style="Accent.TButton",
            command=self.start_conversion
        )
        self.convert_button.pack(
            anchor="center",
            pady=(0, 14)
        )

        progress_area = ttk.Frame(
            action,
            style="Action.TFrame"
        )
        progress_area.pack(
            fill="x",
            padx=(90, 90)
        )

        self.progress = ttk.Progressbar(
            progress_area,
            mode="determinate",
            maximum=100,
            value=0,
            style="Custom.Horizontal.TProgressbar"
        )
        self.progress.pack(
            fill="x"
        )

        self.progress_percent = ttk.Label(
            progress_area,
            text="0%",
            style="Percent.TLabel",
            anchor="center"
        )
        self.progress_percent.pack(
            pady=(7, 0)
        )

        ttk.Label(
            progress_area,
            textvariable=self.main_status,
            style="Status.TLabel",
            anchor="center"
        ).pack(
            fill="x",
            pady=(4, 0)
        )


        self.open_output_button = ttk.Button(
            action,
            text="Open output folder",
            style="Secondary.TButton",
            command=self.open_output_folder,
            state="disabled"
        )
        self.open_output_button.pack(
            anchor="center",
            pady=(12, 0)
        )

    def select_images(self):
        files = filedialog.askopenfilenames(
            title=(
                "Select DJI thermal images"
            ),
            filetypes=[
                (
                    "JPEG images",
                    "*.jpg *.jpeg *.JPG *.JPEG"
                )
            ]
        )

        if not files:
            return

        self.selected_files = [
            Path(path)
            for path in files
        ]

        self.input_folder = None
        self.selection_mode = "files"

        count = len(
            self.selected_files
        )

        if count == 1:
            self.input_path_label.config(
                text=str(
                    self.selected_files[0]
                )
            )

            self.input_status.set(
                "✓ 1 image selected"
            )

        else:
            self.input_path_label.config(
                text=str(
                    self.selected_files[0].parent
                )
            )

            self.input_status.set(
                f"✓ {count} images selected"
            )

        self.input_status_label.configure(
            style="Success.TLabel"
        )

        if self.use_auto_output.get():
            self.output_manually_selected = False

        self.load_source_radiometry(
            self.selected_files[0]
        )

        self.update_auto_output_path()

    def select_folder(self):
        folder = filedialog.askdirectory(
            title=(
                "Select folder with DJI R-JPEG images"
            )
        )

        if not folder:
            return

        folder_path = Path(
            folder
        )

        files = [
            path
            for path in folder_path.iterdir()
            if path.is_file()
            and path.suffix.lower()
            in {".jpg", ".jpeg"}
        ]

        self.selected_files = []
        self.input_folder = folder_path
        self.selection_mode = "folder"

        self.input_path_label.config(
            text=str(folder_path)
        )

        if self.use_auto_output.get():
            self.output_manually_selected = False

        if files:
            self.input_status.set(
                f"✓ Folder selected — "
                f"{len(files)} image(s) found"
            )

            self.input_status_label.configure(
                style="Success.TLabel"
            )

            first_image = sorted(files)[0]
            self.load_source_radiometry(
                first_image
            )

            self.update_auto_output_path()

        else:
            self.source_radiometry = None
            self.clear_radiometry_fields()

            self.input_status.set(
                "No JPG/JPEG images found"
            )

            self.update_auto_output_path()

    def toggle_output_mode(self):
        if self.use_auto_output.get():
            self.output_manually_selected = False
            self.change_output_button.config(
                state="disabled"
            )
            self.update_auto_output_path()

            if self.selection_mode is None:
                self.output_status.set(
                    "✓ Automatic output folder enabled"
                )
            else:
                self.output_status.set(
                    "✓ Automatic output folder selected"
                )

        else:
            self.change_output_button.config(
                state="normal"
            )
            self.output_status.set(
                "Custom output folder mode — choose a location"
            )

    def select_output(self):
        folder = filedialog.askdirectory(
            title="Select output folder",
            initialdir=str(
                self.default_output
            )
        )

        if not folder:
            return

        self.output_path.set(
            folder
        )

        self.output_manually_selected = True

        self.output_status.set(
            "✓ Custom output folder selected"
        )

        self.last_output_dir = Path(
            folder
        )

        self.open_output_button.config(
            state="normal"
        )

    def open_output_folder(self):
        folder = self.last_output_dir

        if folder is None:
            folder = Path(
                self.output_path.get()
            )

        if not folder.exists():
            messagebox.showerror(
                "Output folder",
                (
                    "Output folder does not exist."
                )
            )
            return

        try:
            os.startfile(
                str(folder)
            )

        except Exception as exc:
            messagebox.showerror(
                "Open folder error",
                str(exc)
            )

    def on_radiometry_focus_out(self, variable):
        self.format_radiometry_value(variable)

        if not self.use_image_radiometry.get():
            self.update_auto_output_path()

    @staticmethod
    def format_folder_value(value):
        value = float(value)

        if value.is_integer():
            return str(int(value))

        return (
            f"{value:.4f}"
            .rstrip("0")
            .rstrip(".")
        )

    @staticmethod
    def radiometry_signature(radiometry):
        return (
            round(float(radiometry["emissivity"]), 4),
            round(float(radiometry["distance"]), 4),
            round(float(radiometry["humidity"]), 4),
            round(float(radiometry["reflection"]), 4),
        )

    def get_selected_input_files(self):
        if self.selection_mode == "files":
            return sorted(self.selected_files)

        if (
            self.selection_mode == "folder"
            and self.input_folder is not None
            and self.input_folder.exists()
        ):
            return sorted(
                path
                for path in self.input_folder.iterdir()
                if path.is_file()
                and path.suffix.lower()
                in {".jpg", ".jpeg"}
            )

        return []

    def get_default_output_parent(self):
        if self.selection_mode == "folder":
            return self.input_folder

        if (
            self.selection_mode == "files"
            and self.selected_files
        ):
            return self.selected_files[0].parent

        return None

    def build_parameter_folder_name(self, radiometry):
        emissivity = self.format_folder_value(
            radiometry["emissivity"]
        )
        distance = self.format_folder_value(
            radiometry["distance"]
        )
        humidity = self.format_folder_value(
            radiometry["humidity"]
        )
        reflection = self.format_folder_value(
            radiometry["reflection"]
        )

        return (
            f"TIFF_em_{emissivity}"
            f"_dist_{distance}"
            f"_hum_{humidity}"
            f"_refl_{reflection}"
        )

    def get_source_output_folder_name(self):
        files = self.get_selected_input_files()

        if not files:
            return "TIFF_source_params"

        try:
            sdk = create_sdk()
            signatures = []
            first_radiometry = None

            for image_path in files:
                _, radiometry = sdk.process_image_info(
                    image_path
                )

                if first_radiometry is None:
                    first_radiometry = radiometry

                signatures.append(
                    self.radiometry_signature(
                        radiometry
                    )
                )

            if (
                signatures
                and all(
                    signature == signatures[0]
                    for signature in signatures
                )
            ):
                return self.build_parameter_folder_name(
                    first_radiometry
                )

        except Exception:
            pass

        return "TIFF_source_params"

    def get_custom_output_folder_name(self):
        try:
            radiometry = {
                "emissivity": float(
                    self.emissivity_value
                    .get()
                    .strip()
                    .replace(",", ".")
                ),
                "distance": float(
                    self.distance_value
                    .get()
                    .strip()
                    .replace(",", ".")
                ),
                "humidity": float(
                    self.humidity_value
                    .get()
                    .strip()
                    .replace(",", ".")
                ),
                "reflection": float(
                    self.reflection_value
                    .get()
                    .strip()
                    .replace(",", ".")
                ),
            }

        except (TypeError, ValueError):
            return "TIFF_custom_params"

        return self.build_parameter_folder_name(
            radiometry
        )

    def update_auto_output_path(self):
        if not self.use_auto_output.get():
            return

        if self.output_manually_selected:
            return

        parent = self.get_default_output_parent()

        if parent is None:
            return

        if self.use_image_radiometry.get():
            folder_name = (
                self.get_source_output_folder_name()
            )
        else:
            folder_name = (
                self.get_custom_output_folder_name()
            )

        output_dir = parent / folder_name

        self.output_path.set(
            str(output_dir)
        )

        self.output_status.set(
            "✓ Automatic output folder selected"
        )

        self.last_output_dir = None

        self.open_output_button.config(
            state="disabled"
        )

    def format_radiometry_value(self, variable):
        raw_value = variable.get().strip()

        if not raw_value:
            return

        try:
            value = float(
                raw_value.replace(",", ".")
            )
        except ValueError:
            return

        if value.is_integer():
            variable.set(
                f"{value:.1f}"
            )
        else:
            variable.set(
                str(value)
            )

    def clear_radiometry_fields(self):
        self.emissivity_value.set("")
        self.distance_value.set("")
        self.humidity_value.set("")
        self.reflection_value.set("")

    def apply_source_radiometry(self):
        if not self.source_radiometry:
            self.clear_radiometry_fields()
            return

        self.emissivity_value.set(
            str(self.source_radiometry["emissivity"])
        )
        self.distance_value.set(
            str(self.source_radiometry["distance"])
        )
        self.humidity_value.set(
            str(self.source_radiometry["humidity"])
        )
        self.reflection_value.set(
            str(self.source_radiometry["reflection"])
        )

    def load_source_radiometry(self, image_path):
        try:
            sdk = create_sdk()

            _, radiometry = sdk.process_image_info(
                image_path
            )

            self.source_radiometry = radiometry

            # The displayed values are a preview from the first
            # selected image. During conversion, when the checkbox
            # is enabled, every image still uses its own stored values.
            self.apply_source_radiometry()

        except Exception as exc:
            self.source_radiometry = None
            self.clear_radiometry_fields()

            messagebox.showwarning(
                "Radiometric parameters",
                (
                    "Could not read radiometric parameters "
                    "from the selected image.\n\n"
                    f"{exc}"
                )
            )

    def toggle_radiometry_fields(self):
        use_source = self.use_image_radiometry.get()

        if use_source:
            # Discard edited preview values and restore the
            # values read from the selected source image.
            self.apply_source_radiometry()

        state = (
            "disabled"
            if use_source
            else "normal"
        )

        for entry in self.radiometry_entries:
            entry.configure(
                state=state
            )

        if self.selection_mode is not None:
            self.update_auto_output_path()

    def get_measurement_overrides(self):
        if self.use_image_radiometry.get():
            return None

        field_definitions = [
            {
                "key": "emissivity",
                "label": "Emissivity",
                "raw": self.emissivity_value.get().strip(),
                "minimum": 0.10,
                "maximum": 1.00,
                "range_text": "0.10–1.00",
            },
            {
                "key": "distance",
                "label": "Distance [m]",
                "raw": self.distance_value.get().strip(),
                "minimum": 1.0,
                "maximum": 25.0,
                "range_text": "1–25 m",
            },
            {
                "key": "humidity",
                "label": "Humidity [%]",
                "raw": self.humidity_value.get().strip(),
                "minimum": 1.0,
                "maximum": 100.0,
                "range_text": "1–100%",
            },
            {
                "key": "reflection",
                "label": "Reflected temperature [°C]",
                "raw": self.reflection_value.get().strip(),
                "minimum": -40.0,
                "maximum": 100.0,
                "range_text": "-40–100 °C",
            },
        ]

        values = {}
        errors = []

        for field in field_definitions:
            raw_value = field["raw"]
            label = field["label"]

            if not raw_value:
                errors.append(
                    f"• {label}: value is required "
                    f"(allowed: {field['range_text']})"
                )
                continue

            try:
                value = float(
                    raw_value.replace(",", ".")
                )
            except ValueError:
                errors.append(
                    f"• {label}: '{raw_value}' is not a valid number "
                    f"(allowed: {field['range_text']})"
                )
                continue

            if not (
                field["minimum"]
                <= value
                <= field["maximum"]
            ):
                errors.append(
                    f"• {label}: {raw_value} is outside the allowed range "
                    f"({field['range_text']})"
                )
                continue

            values[field["key"]] = value

        if errors:
            raise ValueError(
                "Please correct the following radiometric parameters:\n\n"
                + "\n".join(errors)
            )

        return values

    def cleanup_old_reports(self, output_dir):
        output_dir = Path(output_dir)

        report_names = (
            "conversion_report.csv",
            "validation_report.csv",
            "DJI_Thermal_Converter_Report.xlsx",
        )

        for report_name in report_names:
            report_path = output_dir / report_name

            if not report_path.exists():
                continue

            try:
                report_path.unlink()
            except OSError:
                # Do not block conversion if an old report is open
                # in Excel or cannot be removed for another reason.
                pass

    def start_conversion(self):
        if self.selection_mode is None:
            messagebox.showerror(
                "No input selected",
                (
                    "Select images or "
                    "an input folder first."
                )
            )
            return

        if self.use_auto_output.get():
            self.update_auto_output_path()

        output_dir = (
            self.output_path
            .get()
            .strip()
        )

        if not output_dir:
            messagebox.showerror(
                "Output folder",
                "Select an output folder."
            )
            return

        try:
            Path(output_dir).mkdir(
                parents=True,
                exist_ok=True
            )

        except OSError as exc:
            messagebox.showerror(
                "Output folder error",
                str(exc)
            )
            return

        self.cleanup_old_reports(Path(output_dir))

        policy = (
            "overwrite"
            if self.overwrite_existing.get()
            else "skip"
        )

        try:
            measurement_overrides = (
                self.get_measurement_overrides()
            )
        except ValueError as exc:
            messagebox.showerror(
                "Radiometric parameters",
                str(exc)
            )
            return

        self.progress.configure(
            value=0
        )

        self.progress_percent.config(
            text="0%"
        )

        self.convert_button.config(
            state="disabled"
        )

        self.open_output_button.config(
            state="disabled"
        )

        self.main_status.set(
            "Starting conversion..."
        )

        thread = threading.Thread(
            target=self.run_conversion,
            args=(
                output_dir,
                policy,
                measurement_overrides
            ),
            daemon=True
        )

        thread.start()

    def run_conversion(
        self,
        output_dir,
        policy,
        measurement_overrides
    ):
        try:
            if self.selection_mode == "files":
                result = convert_images(
                    image_paths=self.selected_files,
                    output_dir=output_dir,
                    progress_callback=(
                        self.progress_callback
                    ),
                    existing_policy=policy,
                    measurement_overrides=measurement_overrides
                )

            else:
                result = convert_folder(
                    input_dir=self.input_folder,
                    output_dir=output_dir,
                    progress_callback=(
                        self.progress_callback
                    ),
                    existing_policy=policy,
                    measurement_overrides=measurement_overrides
                )

            self.root.after(
                0,
                self.set_validation_status
            )

            validation = validate_files(
                result["output_files"]
            )

            error_details = self.read_conversion_errors(
                result["report"]
            )

            self.save_combined_excel_report(
                conversion_report_path=result["report"],
                validation=validation,
                output_dir=output_dir,
                conversion_result=result,
                measurement_overrides=measurement_overrides
            )

            self.root.after(
                0,
                self.conversion_success,
                result,
                validation,
                output_dir,
                error_details
            )

        except Exception as exc:
            self.root.after(
                0,
                self.conversion_error,
                str(exc)
            )

    @staticmethod
    def normalize_report_filename(filename):
        if not filename:
            return ""

        return Path(str(filename)).stem.lower()

    def save_combined_excel_report(
        self,
        conversion_report_path,
        validation,
        output_dir,
        conversion_result,
        measurement_overrides
    ):
        conversion_report_path = Path(
            conversion_report_path
        )
        output_dir = Path(output_dir)

        conversion_rows = []
        conversion_headers = []

        if conversion_report_path.exists():
            with conversion_report_path.open(
                "r",
                newline="",
                encoding="utf-8-sig"
            ) as csv_file:
                reader = csv.DictReader(csv_file)
                conversion_headers = list(
                    reader.fieldnames or []
                )
                conversion_rows = list(reader)

        validation_by_file = {}

        for item in validation.get(
            "results",
            []
        ):
            filename = item.get(
                "filename",
                ""
            )
            key = self.normalize_report_filename(
                filename
            )

            errors = [
                str(value)
                for value in item.get(
                    "errors",
                    []
                )
            ]
            warnings = [
                str(value)
                for value in item.get(
                    "warnings",
                    []
                )
            ]

            status = item.get("status")
            if not status:
                if errors:
                    status = "FAIL"
                elif warnings:
                    status = "WARNING"
                else:
                    status = "PASS"

            validation_by_file[key] = {
                "validation_status": status,
                "validation_errors": " | ".join(errors),
                "validation_warnings": " | ".join(warnings),
            }

        validation_headers = [
            "validation_status",
            "validation_errors",
            "validation_warnings",
        ]

        workbook = Workbook()
        results_sheet = workbook.active
        results_sheet.title = "Results"

        all_headers = (
            conversion_headers
            + [
                header
                for header in validation_headers
                if header not in conversion_headers
            ]
        )

        if not all_headers:
            all_headers = [
                "filename",
                *validation_headers,
            ]

        results_sheet.append(all_headers)

        used_validation_keys = set()

        for conversion_row in conversion_rows:
            filename = (
                conversion_row.get("filename")
                or conversion_row.get("source_filename")
                or conversion_row.get("source")
                or ""
            )
            key = self.normalize_report_filename(
                filename
            )
            validation_row = validation_by_file.get(
                key,
                {}
            )

            if validation_row:
                used_validation_keys.add(key)

            merged_row = dict(conversion_row)
            merged_row.update(validation_row)

            results_sheet.append([
                merged_row.get(header, "")
                for header in all_headers
            ])

        for key, validation_row in validation_by_file.items():
            if key in used_validation_keys:
                continue

            filename = next(
                (
                    item.get("filename", "")
                    for item in validation.get(
                        "results",
                        []
                    )
                    if self.normalize_report_filename(
                        item.get("filename", "")
                    ) == key
                ),
                ""
            )

            merged_row = {
                "filename": filename,
                **validation_row,
            }

            results_sheet.append([
                merged_row.get(header, "")
                for header in all_headers
            ])

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D39A2C"
        )
        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        for cell in results_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        results_sheet.freeze_panes = "A2"
        results_sheet.auto_filter.ref = (
            results_sheet.dimensions
        )

        for column_cells in results_sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
                value = "" if cell.value is None else str(cell.value)
                max_length = max(
                    max_length,
                    len(value)
                )

            results_sheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                40
            )

        summary_sheet = workbook.create_sheet(
            "Summary"
        )

        summary_rows = [
            ["DJI Thermal Converter Report", ""],
            ["Output folder", str(output_dir)],
            ["Total files", conversion_result.get("total", 0)],
            ["Converted", conversion_result.get("success", 0)],
            ["Conversion errors", conversion_result.get("errors", 0)],
            ["Skipped", conversion_result.get("skipped", 0)],
            ["Validation PASS", validation.get("passed", 0)],
            ["Validation WARNING", validation.get("warnings", 0)],
            ["Validation FAIL", validation.get("failed", 0)],
        ]

        if measurement_overrides is None:
            summary_rows.append([
                "Radiometric mode",
                "Values stored in each source image"
            ])
        else:
            summary_rows.extend([
                ["Radiometric mode", "Custom values"],
                ["Emissivity", measurement_overrides.get("emissivity", "")],
                ["Distance [m]", measurement_overrides.get("distance", "")],
                ["Humidity [%]", measurement_overrides.get("humidity", "")],
                ["Reflected temperature [°C]", measurement_overrides.get("reflection", "")],
            ])

        for row in summary_rows:
            summary_sheet.append(row)

        summary_sheet["A1"].font = Font(
            bold=True,
            size=14
        )
        summary_sheet["A1"].fill = header_fill
        summary_sheet["A1"].font = Font(
            bold=True,
            size=14,
            color="FFFFFF"
        )
        summary_sheet.column_dimensions["A"].width = 28
        summary_sheet.column_dimensions["B"].width = 55

        for row in summary_sheet.iter_rows():
            row[0].font = Font(
                bold=True
            )
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        report_path = (
            output_dir
            / "DJI_Thermal_Converter_Report.xlsx"
        )

        workbook.save(report_path)

        # The CSV is still used internally by the converter, but after
        # the final XLSX report is created it is no longer needed by
        # the user. Keep it only if deleting it fails.
        try:
            conversion_report_path.unlink()
        except OSError:
            pass

        return report_path

    def read_conversion_errors(
        self,
        report_path
    ):
        errors = []

        report_path = Path(
            report_path
        )

        if not report_path.exists():
            return errors

        try:
            with report_path.open(
                "r",
                newline="",
                encoding="utf-8-sig"
            ) as csv_file:

                reader = csv.DictReader(
                    csv_file
                )

                for row in reader:
                    if row.get("status") != "ERROR":
                        continue

                    filename = row.get(
                        "filename",
                        "Unknown file"
                    )

                    error = row.get(
                        "error",
                        "Unknown error"
                    )

                    errors.append(
                        f"{filename}\n{error}"
                    )

        except Exception:
            pass

        return errors

    def get_validation_warnings(
        self,
        validation
    ):
        warnings = []

        for result in validation.get(
            "results",
            []
        ):
            filename = result.get(
                "filename",
                "Unknown file"
            )

            file_warnings = result.get(
                "warnings",
                []
            )

            for warning in file_warnings:
                warning_text = str(
                    warning
                )

                if (
                    "Brak opcjonalnego DJI XMP:"
                    in warning_text
                ):
                    field = (
                        warning_text
                        .split(
                            ":",
                            1
                        )[1]
                        .strip()
                    )

                    warning_text = (
                        "missing optional metadata: "
                        f"{field}"
                    )

                elif (
                    "UTCAtExposure"
                    in warning_text
                ):
                    warning_text = (
                        "missing optional metadata: "
                        "UTCAtExposure"
                    )

                warnings.append(
                    f"{filename} — {warning_text}"
                )

        return warnings

    def set_validation_status(self):
        self.main_status.set(
            "Conversion complete — validating TIFF files..."
        )

    def progress_callback(
        self,
        current,
        total,
        success,
        errors,
        skipped
    ):
        if total <= 0:
            percent = 0

        else:
            percent = int(
                (current / total) * 100
            )

        self.root.after(
            0,
            self.update_progress,
            percent,
            current,
            total,
            success,
            errors,
            skipped
        )

    def update_progress(
        self,
        percent,
        current,
        total,
        success,
        errors,
        skipped
    ):
        self.progress.configure(
            value=percent
        )

        self.progress_percent.config(
            text=f"{percent}%"
        )

        status_parts = [
            f"Converted: {current} / {total}"
        ]

        if errors > 0:
            status_parts.append(
                f"Failed: {errors}"
            )

        if skipped > 0:
            status_parts.append(
                f"Skipped: {skipped}"
            )

        self.main_status.set(
            "  •  ".join(
                status_parts
            )
        )

    def conversion_success(
        self,
        result,
        validation,
        output_dir,
        error_details
    ):
        self.last_output_dir = Path(
            output_dir
        )

        self.progress.configure(
            value=100
        )

        self.progress_percent.config(
            text="100%"
        )

        self.convert_button.config(
            state="normal"
        )

        self.open_output_button.config(
            state="normal"
        )

        converted = result["success"]
        total = result["total"]
        skipped = result["skipped"]
        conversion_errors = result["errors"]

        passed = validation["passed"]
        warnings_count = validation["warnings"]
        failed = validation["failed"]

        validation_warnings = (
            self.get_validation_warnings(
                validation
            )
        )

        status_parts = []

        if conversion_errors == 0:
            status_parts.append(
                f"{converted} files converted successfully"
            )
        else:
            status_parts.append(
                f"{converted} of {total} files converted"
            )

        status_parts.append(
            f"{passed} passed all checks"
        )

        if warnings_count > 0:
            status_parts.append(
                f"{warnings_count} passed with warnings"
            )

        if failed > 0:
            status_parts.append(
                f"{failed} failed validation"
            )

        if skipped > 0:
            status_parts.append(
                f"{skipped} skipped"
            )

        self.main_status.set(
            " • ".join(
                status_parts
            )
        )

        message_lines = []

        if (
            conversion_errors == 0
            and failed == 0
        ):
            message_lines.append(
                "Conversion completed successfully"
            )
        else:
            message_lines.append(
                "Conversion completed with issues"
            )

        message_lines.append("")

        message_lines.append(
            f"{converted} of {total} files were converted."
        )

        if skipped > 0:
            message_lines.append(
                f"{skipped} files were skipped."
            )

        if conversion_errors > 0:
            message_lines.append(
                f"{conversion_errors} files could not be converted."
            )

        message_lines.append("")
        message_lines.append(
            "Validation:"
        )

        message_lines.append(
            f"{passed} files passed all checks."
        )

        if warnings_count > 0:
            message_lines.append(
                f"{warnings_count} files are valid "
                "but contain non-critical warnings."
            )

        if failed > 0:
            message_lines.append(
                f"{failed} files failed validation."
            )

        if validation_warnings:
            message_lines.append("")
            message_lines.append(
                "Warnings:"
            )

            for warning in validation_warnings[:5]:
                message_lines.append(
                    warning
                )

            if len(validation_warnings) > 5:
                message_lines.append(
                    f"...and "
                    f"{len(validation_warnings) - 5} more"
                )

        if error_details:
            message_lines.append("")
            message_lines.append(
                "Conversion errors:"
            )

            for error in error_details[:5]:
                message_lines.append("")
                message_lines.append(
                    error
                )

            if len(error_details) > 5:
                message_lines.append("")
                message_lines.append(
                    f"...and "
                    f"{len(error_details) - 5} more"
                )

        message_lines.append("")
        message_lines.append(
            f"Output folder:\n{output_dir}"
        )

        message = "\n".join(
            message_lines
        )

        if (
            conversion_errors == 0
            and failed == 0
        ):
            messagebox.showinfo(
                "Conversion completed",
                message
            )

        else:
            messagebox.showwarning(
                "Conversion completed with issues",
                message
            )

    def conversion_error(
        self,
        error_message
    ):
        self.convert_button.config(
            state="normal"
        )

        self.open_output_button.config(
            state="normal"
        )

        self.main_status.set(
            "Conversion failed"
        )

        messagebox.showerror(
            "Conversion error",
            error_message
        )


def main():
    root = tk.Tk()

    ConverterGUI(
        root
    )

    root.mainloop()


if __name__ == "__main__":
    main()